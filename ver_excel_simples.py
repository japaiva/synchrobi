#!/usr/bin/env python3
"""Ver estrutura do Excel - versão simples"""

try:
    import openpyxl

    excel_path = '/Users/joseantoniopaiva/Downloads/C-Contabeis.xlsx'

    print("=" * 80)
    print(f"ESTRUTURA DO ARQUIVO: C-Contabeis.xlsx")
    print("=" * 80)

    # Abrir workbook
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    print(f"\n📑 Planilhas disponíveis: {wb.sheetnames}")

    # Pegar primeira planilha
    ws = wb[wb.sheetnames[0]]

    print(f"\n📊 Planilha ativa: {ws.title}")
    print(f"   Dimensões: {ws.max_row} linhas x {ws.max_column} colunas")

    # Ler cabeçalhos (primeira linha)
    print(f"\n📋 Colunas (primeira linha):")
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        headers.append(cell_value)
        print(f"   {col-1}: {cell_value}")

    # Ler primeiras 5 linhas de dados
    print(f"\n👀 Primeiras 5 linhas de dados:")
    for row in range(2, min(7, ws.max_row + 1)):
        print(f"\n   Linha {row}:")
        for col in range(1, min(ws.max_column + 1, 10)):  # Primeiras 10 colunas
            header = headers[col-1] if col-1 < len(headers) else f"Col{col}"
            value = ws.cell(row, col).value
            if value:
                print(f"      {header}: {value}")

    wb.close()

except ImportError:
    print("❌ openpyxl não está instalado")
    print("Tente: pip install openpyxl")
except Exception as e:
    print(f"❌ Erro: {e}")
    import traceback
    traceback.print_exc()

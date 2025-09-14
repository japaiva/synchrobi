# gestor/views/movimento.py - VERSÃO LIMPA SEM IMPORTAÇÃO

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.utils import timezone
from datetime import datetime, date, timedelta
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io

from core.models import Movimento, Unidade, CentroCusto, ContaContabil, ContaExterna, Fornecedor
from core.forms import MovimentoForm

logger = logging.getLogger('synchrobi')

# === VIEWS PRINCIPAIS ===

@login_required
def movimento_list(request):
    """Lista de movimentos com filtros"""
    search = request.GET.get('search', '')
    ano = request.GET.get('ano', '')
    mes = request.GET.get('mes', '')
    unidade = request.GET.get('unidade', '')
    centro_custo = request.GET.get('centro_custo', '')
    
    movimentos = Movimento.objects.select_related(
        'unidade', 'centro_custo', 'conta_contabil', 'fornecedor'
    ).order_by('-data', '-id')
    
    if search:
        movimentos = movimentos.filter(
            Q(historico__icontains=search) |
            Q(documento__icontains=search) |
            Q(fornecedor__razao_social__icontains=search) |
            Q(fornecedor__codigo__icontains=search)
        )
    
    if ano:
        movimentos = movimentos.filter(ano=int(ano))
    
    if mes:
        movimentos = movimentos.filter(mes=int(mes))
    
    if unidade:
        movimentos = movimentos.filter(unidade_id=unidade)
    
    if centro_custo:
        movimentos = movimentos.filter(centro_custo__codigo__icontains=centro_custo)
    
    # Calcular totais
    total_movimentos = movimentos.count()
    total_valor = movimentos.aggregate(total=Sum('valor'))['total'] or 0
    
    # Paginação
    paginator = Paginator(movimentos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Preparar dados para os dropdowns
    anos_disponiveis = list(Movimento.objects.values_list('ano', flat=True).distinct().order_by('-ano'))
    unidades_disponiveis = Unidade.objects.filter(ativa=True).order_by('codigo')
    
    meses = [
        (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
        (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
        (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro')
    ]
    
    context = {
        'page_obj': page_obj,
        'total_movimentos': total_movimentos,
        'total_valor': total_valor,
        'search': search,
        'anos_disponiveis': anos_disponiveis,
        'unidades_disponiveis': unidades_disponiveis,
        'meses': meses,
        'filtros': {
            'ano': ano,
            'mes': mes,
            'unidade_id': unidade,
            'centro_custo': centro_custo,
            'search': search,
        }
    }
    
    return render(request, 'gestor/movimento_list.html', context)

@login_required
def movimento_create(request):
    """Criar novo movimento"""
    if request.method == 'POST':
        form = MovimentoForm(request.POST)
        if form.is_valid():
            try:
                movimento = form.save()
                messages.success(request, f'Movimento criado com sucesso! Valor: {movimento.valor_formatado}')
                logger.info(f'Movimento criado: {movimento.id} - {movimento.valor_formatado} por {request.user}')
                return redirect('gestor:movimento_list')
            except Exception as e:
                messages.error(request, f'Erro ao criar movimento: {str(e)}')
                logger.error(f'Erro ao criar movimento: {str(e)}')
        else:
            messages.error(request, 'Erro ao criar movimento. Verifique os dados.')
    else:
        form = MovimentoForm()
        
        # Pré-popular com mês/ano atual se fornecido
        mes_inicial = request.GET.get('mes')
        ano_inicial = request.GET.get('ano')
        if mes_inicial:
            form.fields['mes'].initial = mes_inicial
        if ano_inicial:
            form.fields['ano'].initial = ano_inicial
    
    context = {
        'form': form,
        'title': 'Novo Movimento',
        'is_create': True
    }
    return render(request, 'gestor/movimento_form.html', context)

@login_required
def movimento_update(request, pk):
    """Editar movimento"""
    movimento = get_object_or_404(Movimento, pk=pk)
    
    # Guardar valores originais para log
    valores_originais = {
        'valor': movimento.valor,
        'historico': movimento.historico,
        'fornecedor': movimento.fornecedor,
        'natureza': movimento.natureza
    }
    
    if request.method == 'POST':
        form = MovimentoForm(request.POST, instance=movimento)
        if form.is_valid():
            try:
                movimento_atualizado = form.save()
                
                # Log de alterações
                alteracoes = []
                for campo, valor_original in valores_originais.items():
                    valor_novo = getattr(movimento_atualizado, campo)
                    if valor_original != valor_novo:
                        alteracoes.append(f"{campo}: {valor_original} → {valor_novo}")
                
                if alteracoes:
                    logger.info(f'Movimento {movimento.id} alterado por {request.user}: {", ".join(alteracoes)}')
                
                messages.success(request, f'Movimento atualizado com sucesso! Valor: {movimento_atualizado.valor_formatado}')
                return redirect('gestor:movimento_list')
            except Exception as e:
                messages.error(request, f'Erro ao atualizar movimento: {str(e)}')
                logger.error(f'Erro ao atualizar movimento {movimento.id}: {str(e)}')
        else:
            messages.error(request, 'Erro ao atualizar movimento. Verifique os dados.')
    else:
        form = MovimentoForm(instance=movimento)
    
    context = {
        'form': form,
        'title': 'Editar Movimento',
        'movimento': movimento,
        'is_create': False
    }
    return render(request, 'gestor/movimento_form.html', context)

@login_required
def movimento_delete(request, pk):
    """Deletar movimento"""
    movimento = get_object_or_404(Movimento, pk=pk)
    
    if request.method == 'POST':
        periodo_display = movimento.periodo_display
        valor_formatado = movimento.valor_formatado
        historico_resumido = movimento.historico[:50]
        movimento_id = movimento.id
        
        try:
            movimento.delete()
            messages.success(
                request, 
                f'Movimento excluído com sucesso! '
                f'Período: {periodo_display}, Valor: {valor_formatado}'
            )
            logger.info(f'Movimento excluído: {movimento_id} - {valor_formatado} por {request.user}')
            return redirect('gestor:movimento_list')
        except Exception as e:
            messages.error(request, f'Erro ao excluir movimento: {str(e)}')
            logger.error(f'Erro ao excluir movimento {movimento_id}: {str(e)}')
            return redirect('gestor:movimento_list')
    
    context = {
        'movimento': movimento
    }
    return render(request, 'gestor/movimento_delete.html', context)

@login_required
def movimento_export_excel(request):
    """Exportar movimentos para Excel"""
    
    try:
        # Aplicar os mesmos filtros da listagem
        search = request.GET.get('search', '')
        ano = request.GET.get('ano', '')
        mes = request.GET.get('mes', '')
        unidade = request.GET.get('unidade', '')
        centro_custo = request.GET.get('centro_custo', '')
        
        movimentos = Movimento.objects.select_related(
            'unidade', 'centro_custo', 'conta_contabil', 'fornecedor'
        ).order_by('-data', '-id')
        
        # Aplicar filtros
        if search:
            movimentos = movimentos.filter(
                Q(historico__icontains=search) |
                Q(documento__icontains=search) |
                Q(fornecedor__razao_social__icontains=search) |
                Q(fornecedor__codigo__icontains=search)
            )
        
        if ano:
            movimentos = movimentos.filter(ano=int(ano))
        
        if mes:
            movimentos = movimentos.filter(mes=int(mes))
        
        if unidade:
            movimentos = movimentos.filter(unidade_id=unidade)
        
        if centro_custo:
            movimentos = movimentos.filter(centro_custo__codigo__icontains=centro_custo)
        
        # Limitar exportação para evitar problemas de memória
        total_movimentos = movimentos.count()
        if total_movimentos > 50000:
            messages.warning(
                request, 
                f'Muitos registros para exportar ({total_movimentos:,}). '
                f'Use filtros para reduzir o volume ou exporte por partes.'
            )
            return redirect('gestor:movimento_list')
        
        # Criar workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimentos"
        
        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Cabeçalhos
        headers = [
            'Data', 'Mês', 'Ano', 'Período',
            'Código Unidade', 'Nome Unidade', 'Código All Strategy',
            'Código Centro Custo', 'Nome Centro Custo',
            'Código Conta Contábil', 'Nome Conta Contábil',
            'Código Fornecedor', 'Razão Social Fornecedor',
            'Documento', 'Natureza', 'Valor', 'Histórico',
            'Código Projeto', 'Gerador', 'Rateio',
            'Data Importação', 'Arquivo Origem', 'Linha Origem'
        ]
        
        # Escrever cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_thin
        
        # Escrever dados
        for row, movimento in enumerate(movimentos, 2):
            data_row = [
                movimento.data.strftime('%d/%m/%Y') if movimento.data else '',
                movimento.mes,
                movimento.ano,
                movimento.periodo_display,
                movimento.unidade.codigo if movimento.unidade else '',
                movimento.unidade.nome if movimento.unidade else '',
                movimento.unidade.codigo_allstrategy if movimento.unidade else '',
                movimento.centro_custo.codigo if movimento.centro_custo else '',
                movimento.centro_custo.nome if movimento.centro_custo else '',
                movimento.conta_contabil.codigo if movimento.conta_contabil else '',
                movimento.conta_contabil.nome if movimento.conta_contabil else '',
                movimento.fornecedor.codigo if movimento.fornecedor else '',
                movimento.fornecedor.razao_social if movimento.fornecedor else '',
                movimento.documento,
                movimento.natureza,
                float(movimento.valor) if movimento.valor else 0,
                movimento.historico,
                movimento.codigo_projeto,
                movimento.gerador,
                movimento.rateio,
                movimento.data_importacao.strftime('%d/%m/%Y %H:%M') if movimento.data_importacao else '',
                movimento.arquivo_origem,
                movimento.linha_origem
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border_thin
                
                # Formatação especial para valores monetários
                if col == 16:  # Coluna Valor
                    cell.number_format = '#,##0.00'
                    if value < 0:
                        cell.font = Font(color="FF0000")  # Vermelho para negativos
        
        # Ajustar largura das colunas
        column_widths = [
            12, 8, 8, 10,  # Data, Mês, Ano, Período
            15, 30, 15,     # Unidade
            15, 30,         # Centro Custo
            15, 30,         # Conta Contábil
            15, 40,         # Fornecedor
            15, 8, 15,      # Documento, Natureza, Valor
            50,             # Histórico
            15, 15, 8,      # Projeto, Gerador, Rateio
            20, 25, 8       # Importação info
        ]
        
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # Adicionar aba de resumo
        ws_resumo = wb.create_sheet(title="Resumo")
        
        # Estatísticas para resumo
        stats = {
            'Total de Movimentos': total_movimentos,
            'Soma de Valores': float(movimentos.aggregate(Sum('valor'))['valor__sum'] or 0),
            'Movimentos Débito': movimentos.filter(natureza='D').count(),
            'Movimentos Crédito': movimentos.filter(natureza='C').count(),
            'Fornecedores Únicos': movimentos.exclude(fornecedor__isnull=True).values('fornecedor').distinct().count(),
            'Unidades Únicas': movimentos.values('unidade').distinct().count(),
            'Data de Exportação': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
            'Exportado por': request.user.get_full_name() or request.user.username
        }
        
        # Escrever resumo
        for row, (label, value) in enumerate(stats.items(), 1):
            ws_resumo.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws_resumo.cell(row=row, column=2, value=value)
        
        # Salvar em memory
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Criar resposta HTTP
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nome do arquivo baseado nos filtros
        filename_parts = ['movimentos']
        if ano:
            filename_parts.append(str(ano))
        if mes:
            filename_parts.append(f"mes_{int(mes):02d}")
        
        filename = '_'.join(filename_parts) + '_' + timezone.now().strftime('%Y%m%d_%H%M%S') + '.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f'Exportação Excel realizada por {request.user}: {total_movimentos} movimentos')
        messages.success(request, f'Exportação concluída! {total_movimentos:,} movimentos exportados.')
        
        return response
        
    except Exception as e:
        logger.error(f'Erro na exportação Excel: {str(e)}')
        messages.error(request, f'Erro na exportação: {str(e)}')
        return redirect('gestor:movimento_list')